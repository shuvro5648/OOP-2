import openpyxl


def create():
    excel_sheet = openpyxl.Workbook()
    sheet = excel_sheet.active
    sheet['A1'] = 'Day'
    sheet['B1'] = 'Expense'
    sheet['C1'] = 'Daily Budget'
    sheet['D1'] = 'Total Budget'

    excel_sheet.save('Excel.xlsx')


# def load_Workbook(param):
#     pass


def main():
    create()
    total_budget = float(input("##Enter the total budget: $"))
    daily_budget = float(input("##Enter the daily budget: $"))
    total_expense = 0
    extended_days = []
    week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for days in week_days:
        daily_expense = float(input(f"Enter Expense for {days}: $"))
        total_expense = daily_expense

        if daily_expense > daily_budget:
            extended_days.append((days, daily_expense))
        excel_sheet = openpyxl.load_workbook('Excel.xlsx')
        sheet = excel_sheet.active
        sheet.append([days, daily_expense])
        excel_sheet.save('Excel.xlsx')
    daily_avg_expense = total_expense/7
    daily_avg_expense = format(daily_avg_expense, ".2f")
    print("Expense Summery: ")
    print(f"Total Expenses for the week: ${total_expense}")
    print(f"Average Daily Expense: ${daily_avg_expense}")

    if extended_days:
        print("Days where expenses exceeded the daily budget: ")
        for day, expense in extended_days:
            print(f"{day}: ${expense}")




main()
